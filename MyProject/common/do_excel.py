# -*-coding:utf-8-*-

"""
@Time       :2019/2/1811:39
@Author     :pengpeng
@Email      :743463927@qq.com
@File       :do_excel.py
@Software   :PyCharm
@Function   :
"""

import openpyxl
from common import logger

logger = logger.get_logger('do_excel')


class Case:
    '''
    测试用例封装类
    '''

    def __init__(self):
        self.id = None
        self.url = None
        self.data = None
        self.title = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None


class DoExcel:

    def __init__(self, file_name):
        try:
            # 操作的文件
            self.file_name = file_name
            # 实例化一个workbook对象
            self.workbook = openpyxl.load_workbook(filename=self.file_name)
            # 异常处理
        except FileNotFoundError as e:
            logger.error('{0} not found,please check file path'.format(file_name))
            raise e

    def get_cases(self, sheet_name):
        sheet = self.workbook[sheet_name]  # 获取sheet
        max_row = sheet.max_row  # 获取sheet最大行数
        cases = []  # 定义一个列表，用来存放即将要放进去的测试用例
        for r in range(2, max_row + 1):  # for 循环，从第二行开始遍历
            case = Case()  # 实例化一个case对象，用来存放测试数据
            case.id = sheet.cell(row=r, column=1).value  # 取第r行，第1格的数据
            case.title = sheet.cell(row=r, column=2).value
            case.url = sheet.cell(row=r, column=3).value
            case.data = sheet.cell(row=r, column=4).value
            case.method = sheet.cell(row=r, column=5).value
            case.expected = sheet.cell(row=r, column=6).value
            if type(case.expected) == int:
                case.expected = str(case.expected)
            cases.append(case)  # 将case放到cases列表里面
        return cases  # for 循环结束后返回cases列表

    def write_result(self, sheet_name, row, actual, result):
        sheet = self.workbook[sheet_name]  # 获取sheet
        sheet.cell(row, 7).value = actual  # 写入实际结果
        sheet.cell(row, 8).value = result  # 写入执行结果，pass or fail
        self.workbook.save(filename=self.file_name)
